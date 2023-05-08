# classmethod and staticmethod

# Functions and arguments
# functions are defined with 'def' and must be ended with a column ':' just before '()'
# then followed by the name if the function
# And a set of formal parameters
# actual parameters or arguments are passed during a function call
# (i.e) actual parameters and arguments when the function is called
###
# functions can be defined with variable number of arguments.

# Example:
# def realtors(arg1, arg2)
#   functions statements are written

#   realtors(a,b)

###
# def -- defines realtors as to be a function
# arg1 and arg2 are the formal parameters

# realtors(a,b) here the function is called
# (a, b) are the actual parameter or arguments

###

# Function types

# DEFAULT ARGUMENTS


def add(a, b=5, c=10):
    return a+b+c


'''

Default arguments are values that are provided while defining functions.
The assignment operator = is used to assign a default value to the argument.
Default arguments become optional during the function calls.
If we provide a value to the default arguments during function calls, it overrides the default value.
The function can have any number of default arguments.
Default arguments should follow non-default arguments.

'''
###
# GIVING ONLY THE MANDATORY ARGUMENT
print(add(3))

# GIVING ONLY THE OPTIONAL ARGUMENTS
print(add(3, 4))

# GIVING ALL THE MANDATORY ARGUMENT
print(add(2, 3, 4))


###
# functions can be called by using keyword arguments
# that is keyword=value (of keyword)
# see below


def add(a, b=5, c=10):
    return a + b + c


# there is no need to maintain the same order is the formal argument passed
# the keyword parameter makes this easier
# 'b' is the keyword and "10" the value of the keyword

print(add(b=10, c=15, a=20))

# note the function add has an argument that has no keyword value which is 'a'
# which mean it can be assigned a value when the add function is call
# this is the DEFAULT optional ARGUMENT type
print(add(a=10))


# output will be 25 (i.e) a=10 + b=5 + c=10

###
# POSITIONAL ARGUMENT
# During a function call
# Values passed through arguments should be in the order of parameters in the defined function
# (i.e) THE SAME WAY IT PASSED IN THE FUNCTION IS THE SAME WAY IT HAS TO BE WHEN CALLING FUNCTION
# EXAMPLE:


def add(a, b, c):
    return a + b + c


# values passed in the argument are passed to parameters or FORMAL ARGUMENTS by their position
print(add(10, 20, 30))
# note a=10 b=20 c=30
# finally keyword arguments do not require same order of argument positioning as to positional arguments

###

# Variable-length Arguments
###
'''
Variable-length arguments are also known as arbitrary arguments. 
If we don’t know the number of arguments needed for the function in advance, 
we can use arbitrary arguments
'''

###

'''
THERE ARE TWO TYPES OF ARBITRARY ARGUMENTS:
Arbitrary positional arguments.
Arbitrary keyword arguments.


ARBITRARY POSITIONAL ARGUMENTS IN PYTHON
For arbitrary positional argument, 
an asterisk (*) is placed before a parameter in function definition 
"WHICH CAN HOLD NON-KEYWORD VARIABLE-LENGTH ARGUMENTS."
These arguments will be wrapped up in a tuple. Before the variable number of arguments, 
zero or more normal arguments may occur.
'''


# Example:


def add(*b):
    result = 0
    for i in b:
        result = result + i
    return result


print(add(1, 2, 3, 4, 5))

print(add(10, 20))

"""
ARBITRARY KEYWORD ARGUMENTS IN PYTHON
For arbitrary positional argument, 
A DOUBLE ASTERISK (**) IS PLACED BEFORE A PARAMETER IN A FUNCTION 
which can hold keyword variable-length arguments.
"""


def fn(**a):
    for i in a.items():
        print(i)


fn(numbers=5, colors="blue", fruits="apple")

# Output:
# ('numbers', 5)
# ('colors', 'blue')
# ('fruits', 'apple')



"""
Python crash course

Twitter @i_amgoke: https://twitter.com/i_amgoke

Github: https://www.github.com/Gokegold

Date Created: April 28, 2023

Date modified: May 8, 2023

"""
import openpyxl as xl
from openpyxl.chart import BarChart, Reference

from nome import Employee

'''
print('Hello World')


class Point:
    def __init__(self, x, y):
        self.x = x
        self.y = y

    def move(self):
        print("move")

    def draw(self):
        print("draw")


point = Point(10, 20)
point.x = 12
print(point.x)


class Person:
    def __init__(self, first_name, last_name):
        self.first_name = first_name
        self.last_name = last_name
    def talk(self):
        print(f"Hi, I am {self.first_name}, {self.last_name}")


john = Person("John", "Smith")
john.talk()

bob = Person("Bob", "Smith")
bob.talk()Bible Book List
Font Size


sudan = Person("Sudan", "Smith")
print(sudan.first_name, sudan.last_name)
sudan.talk()


import nome

print(f"Welcome On Board")

first_name = input("First Name: ")
last_name = input("Last Nam: ")
age = int(input("DOB: "))
gender = input("Gender: ")

print(nome.ClientsDetails)



first_name = "John"
last_name = "Sarah"


def greet():
    return f"Congratulations Dear", first_name, last_name


strings = greet()
print(strings)

n = 300
m = n

print(m)


cell = sheet['a1']
sheet.cell(1, 1)
print(cell.value)
print(sheet.max_row)
def process_workbook(filename):
    wb = xl.load_workbook(filename)
    sheet = wb['Sheet1']

    for row in range(2, sheet.max_row + 1):
        cell = sheet.cell(row, 3)
        corrected_price = cell.value * 0.9
        corrected_price_cell = sheet.cell(row, 4)
        corrected_price_cell.value = corrected_price

    values = Reference(
            min_row=2,
            max_row=sheet.max_row,
            min_col=4,
            max_col=4)

    chart = BarChart()
    chart.add_data(values)
    sheet.add_chart(chart, 'e2')
    wb.save(filename)

'''
# classmethod and staticmethod

# Functions and arguments
# functions are defined with 'def' and must be ended with a column ':' just before '()'
# then followed by the name if the function
# And a set of formal parameters
# actual parameters or arguments are passed during a function call
# (i.e) actual parameters and arguments when the function is called
###
# functions can be defined with variable number of arguments.

# Example:
# def realtors(arg1, arg2)
#   functions statements are written

#   realtors(a,b)

###
# def -- defines realtors as to be a function
# arg1 and arg2 are the formal parameters

# realtors(a,b) here the function is called
# (a, b) are the actual parameter or arguments

###

# Function types

# DEFAULT ARGUMENTS


def add(a, b=5, c=10):
    return a+b+c


'''

Default arguments are values that are provided while defining functions.
The assignment operator = is used to assign a default value to the argument.
Default arguments become optional during the function calls.
If we provide a value to the default arguments during function calls, it overrides the default value.
The function can have any number of default arguments.
Default arguments should follow non-default arguments.

'''
###
# GIVING ONLY THE MANDATORY ARGUMENT
print(add(3))

# GIVING ONLY THE OPTIONAL ARGUMENTS
print(add(3, 4))

# GIVING ALL THE MANDATORY ARGUMENT
print(add(2, 3, 4))


###
# functions can be called by using keyword arguments
# that is keyword=value (of keyword)
# see below


def add(a, b=5, c=10):
    return a + b + c


# there is no need to maintain the same order is the formal argument passed
# the keyword parameter makes this easier
# 'b' is the keyword and "10" the value of the keyword

print(add(b=10, c=15, a=20))

# note the function add has an argument that has no keyword value which is 'a'
# which mean it can be assigned a value when the add function is call
# this is the DEFAULT optional ARGUMENT type
print(add(a=10))


# output will be 25 (i.e) a=10 + b=5 + c=10

###
# POSITIONAL ARGUMENT
# During a function call
# Values passed through arguments should be in the order of parameters in the defined function
# (i.e) THE SAME WAY IT PASSED IN THE FUNCTION IS THE SAME WAY IT HAS TO BE WHEN CALLING FUNCTION
# EXAMPLE:


def add(a, b, c):
    return a + b + c


# values passed in the argument are passed to parameters or FORMAL ARGUMENTS by their position
print(add(10, 20, 30))
# note a=10 b=20 c=30
# finally keyword arguments do not require same order of argument positioning as to positional arguments

###

# Variable-length Arguments
###
'''
Variable-length arguments are also known as arbitrary arguments. 
If we don’t know the number of arguments needed for the function in advance, 
we can use arbitrary arguments
'''

###

'''
THERE ARE TWO TYPES OF ARBITRARY ARGUMENTS:
Arbitrary positional arguments.
Arbitrary keyword arguments.


ARBITRARY POSITIONAL ARGUMENTS IN PYTHON
For arbitrary positional argument, 
an asterisk (*) is placed before a parameter in function definition 
"WHICH CAN HOLD NON-KEYWORD VARIABLE-LENGTH ARGUMENTS."
These arguments will be wrapped up in a tuple. Before the variable number of arguments, 
zero or more normal arguments may occur.
'''


# Example:


def add(*b):
    result = 0
    for i in b:
        result = result + i
    return result


print(add(1, 2, 3, 4, 5))

print(add(10, 20))

"""
ARBITRARY KEYWORD ARGUMENTS IN PYTHON
For arbitrary positional argument, 
A DOUBLE ASTERISK (**) IS PLACED BEFORE A PARAMETER IN A FUNCTION 
which can hold keyword variable-length arguments.
"""


def fn(**a):
    for i in a.items():
        print(i)


fn(numbers=5, colors="blue", fruits="apple")

# Output:
# ('numbers', 5)
# ('colors', 'blue')
# ('fruits', 'apple')


print(Employee.raise_amt())
print(emp_1.raise_amt())
print(emp_2.raise_amt)

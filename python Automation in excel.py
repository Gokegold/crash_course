# The purpose is to learn how to interact with excel using python.
# By storing a new value in spcific new rows of new columns

import openpyxl as xl
from openpyxl.chart import BarChart, Reference
wb = xl.load_workbook('transactions.xlsx')
sheet = wb['Sheet1']

for row in range(2, sheet.max_row, + 1):
# The row and column to be used is declared in the cell variable
  cell = sheet.cell(row, 3)
# The correction is assigned to the variable correct_price
  corrected_price = cell.value * 0.9
# Then the row to assign the corrected_price is stated and assigned to corrected_price_cell variable
  corrected_price_cell = sheet.cell(row, 4)
  corrected_price_cell.value. corrected_price

  
  # Now declare reference and set coordinate
  # Then asign it to the varaible values
  
  values =reference(
    main_row = 2
    max_row =sheet.max_row,
    min_col = 4
    max_col = 4
  )

# Declaring what type of is to be used.
chart = BarChart
# Adding the data to the chart
chart.add_data(values)
# Adding the chart to the sheet
sheet.add.chart(chart, 'e2')


wb.save('transaction2.xlsx')


def process_workbook(filename):
    wb = xl.load_workbook(filename)
    sheet = wb['Sheet1']

    for row in range(2, sheet.max_row + 1):
        cell = sheet.cell(row, 3)
        corrected_price = cell.value * 0.9
        corrected_price_cell = sheet.cell(row, 4)
        corrected_price_cell.value = corrected_price

    values = Reference(
            min_row=2,
            max_row=sheet.max_row,
            min_col=4,
            max_col=4)

    chart = BarChart()
    chart.add_data(values)
    sheet.add_chart(chart, 'e2')
    wb.save(filename)

